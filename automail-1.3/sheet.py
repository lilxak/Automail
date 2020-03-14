import openpyxl as op

def sheet_to_dict(path):
    ''' Inputs an excel workbook's path and returns a dictionary {'Column':[Rows]} '''

    name = op.load_workbook(path).sheetnames[0]
    sheet = op.load_workbook(path)[name]
    output = {}
    col_names = [a.value for a in sheet[1]]
    for c in col_names:
            output[c] = []
    
    cols = ['A','B','C','D','E','F','G','H','I','J']
    for c,k in zip(cols, col_names):
        new = [col.value for col in sheet[c]]
        output[k] = new[1:]
    if None in output.keys():
        del output[None]
    
    return output